# -*- coding: utf-8 -*-
"""
PDMAC
Pipeline Defect Modeling and Analysis Calculator
Created on 08/18/2022
Updated on 04/10/2024

@author: evalencia
"""

# =============================================================================
# ONLY CHANGE THESE VALUES
# =============================================================================

# Adjust to gas or liquid depending on product type
data_type   = 'neither'

# Select the ILI vendor type
# ili_format  = 'entegra'
ili_format  = 'southern'

# If performing analysis on liquid pipelines, only need to observe the START and END pressure history. 
# Will need to remove any null/empty/nan values that are found in either the START and END. If found in one, remove from both.

# Data structure should be:
#   time_col    Time stamp, data in second intervals
#   P_cols      Index of Pressure Columns, data in psi
time_col = 0
P_cols = [1]

# Leave smoothing as True unless for special circumstances
smoothing = True

# =============================================================================
# DO NOT EDIT BELOW
# =============================================================================

# import numpy as np
import pandas as pd
# import tkinter as tk
# import rainflow
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import time
from datetime import datetime
import os
# Custom packages
import rainflow_analysis
import dent_process
import dent_strain
# import remaining_life
# import reports

time_start = time.time()

# =============================================================================
# Controls
# =============================================================================

wbR_path = 'results/all_results.xlsx'
time_ref = '%03d | '
press_hist_skiprows = 1
press_hist_header   = None

# PDMAC SCF Calculations
circ_int = 0.25  # Circumferential interval length, in
axial_int = 0.25 # Axial interval length, in

# Graphing
plt.rcParams['font.family'] = 'Times New Roman'
plt.rcParams['mathtext.fontset'] = 'cm'
plt.rcParams['font.size'] = 8
plt.rcParams['lines.markersize'] = 0.5

# =============================================================================
# Iterate through Data Files
# =============================================================================

# Import the dent index dataframe
df_index = pd.read_excel('index.xlsx', header=1)

# Create an output dataframe for the results
results_columns = pd.read_excel('results/all_results.xlsx', header=1).columns
df_results = pd.DataFrame(columns=results_columns)
results_data = []
j = 0

# Iterate through all of the dent files
for subdir, dirs, files in os.walk('data'):

    # Skip the first iteration of each new folder
    if (not '\\' in subdir) or (len(files) == 0):
        continue
    
    # Ignore folder titled 'ignore'
    if 'ignore' in subdir.lower():
        continue
        
    # Keep track of time for each subdir folder
    time_subdir_start = time.time()
    print('========== SUBDIR START ==========')
    print((time_ref + 'Processing Sub Directory: %s') % (time.time() - time_start, subdir))
    
    # Perform Rainflow Analysis and Dent Strain for every dent within this pipeline segment
    df_segment_dents = df_index.loc[df_index['Pipeline Segment'].astype('string') == subdir.split('\\')[1]]
    if df_segment_dents.size == 0:
        print((time_ref + 'No dents found in Sub Directory: %s') % (time.time() - time_start, subdir))
        continue
    
    # Find the pressure history for the corresponding pipeline segment
    print((time_ref + 'Analyzing the pressure history for the pipeline segment...') % (time.time() - time_start))
    press_hist_path = os.path.join(subdir, 'pressure_history.xlsx')
    # If the product is liquid, only the P_list will be returned.
    if data_type == 'liquid':
        P_list, P_time = rainflow_analysis.liquid_P(press_hist_path, P_cols, press_hist_header, press_hist_skiprows, time_col)
    elif data_type == 'gas':
        # Save the pressure history in the Results folder but with the name of the segment
        graph_path = 'results/' + subdir.split('\\')[1]
        # Collect a sample dent from the segment, need to use OD, WT, SMYS, Service Years, Min Range, and M
        SSI, CI, SSI_MD49, cycles, MD49_bins = rainflow_analysis.gas(press_hist_path, graph_path, P_cols, press_hist_header, press_hist_skiprows, df_segment_dents.iloc[0], time_col)
    else:
        print((time_ref + 'Not analyzing pressure data.') % (time.time() - time_start))
        SSI = ''
        CI = ''
        SSI_MD49 = ''
        P_time = ''
        
    print((time_ref + 'Done.') % (time.time() - time_start))
                
    for i in range(len(df_segment_dents)):
        # Extract the information for the specified dent
        dent_ID, pipe_segment, odometer, OD, WT, SMYS, service_years, min_range, M, fatigue_curve, dent_depth, dent_depth_in, dent_length, dent_width, dent_orientation, Lx, hx, SG, K, L1, L2, h1, h2, D1, D2, flag, *_ = df_segment_dents.iloc[i]
        
        # Find a matching str in files. There should only be one match.
        file_name = [file_name for file_name in files if str(dent_ID) in file_name]
        if file_name == []:
            continue
        
        # Skip features that were previously repaired
        if flag == 'repaired' or flag == 'ignore':
            print((time_ref + 'Dent ID %s was repaired. Continuing to next feature.') % (time.time() - time_start, dent_ID))
            continue
        
        print('=========== FILE START ===========')
        print((time_ref + 'Processing Dent ID: %s') % (time.time() - time_start, dent_ID))
        
        # Create an output folder for the dent analysis results.
        results_path = 'results/' + str(dent_ID)
        try:
            os.makedirs(results_path, exist_ok=False)
        except:
            continue
        results_path = results_path + '/'
        
        # Rainflow Analysis if liquid product, it needs to run with individual dent files
        if data_type == 'liquid':
            SSI, CI, SSI_MD49, cycles, MD49_bins = rainflow_analysis.liquid(P_list, P_time, results_path, df_segment_dents.iloc[i])
            print((time_ref + 'Rainflow Analysis results are done.') % (time.time() - time_start))

        # PDMAC SCF Calculations
        dent_path = os.path.join(subdir, file_name[0])
        # ADDED LAST TWO PARAMETERS 6/28 TESTING SMOOTHING 
        SCF, MPs, df_data, flag = dent_process.process_dent(dent_ID, dent_path, results_path, ili_format, OD, WT, SMYS, circ_int, axial_int, smoothing, time_start)
        
        # Export the Smooth Data for use in MD49.
        
        # Generate a report based on the SCF reports
        int_review_path = results_path + 'Abaqus Results/Internal Review/'
        interaction = 'N/A'
        # reports.SCF(results_path, int_review_path, dent_ID, interaction, OD, WT, dent_depth_in, dent_length, dent_width, MPs, SCF)
        
        # Dent Strain Calculations
        eo_max, ei_max, strain_limit, strain_pass_fail = dent_strain.strain(dent_ID, results_path, OD, WT, dent_depth_in, dent_length, df_data, circ_int, axial_int, time_start)
        
        # Remaining Life using SCF
        # remaining_life.export(dent_ID, results_path, OD, WT, SMYS, service_years, min_range, M, cycles, MD49_bins, SCF, fatigue_curve)
        
        remaining_life_val = 0
        
        # Create a list of all of the results to save in the results summary
        results = [str(datetime.fromtimestamp(time.time())),
                    dent_ID,
                    # Pipeline Information
                    pipe_segment,
                    odometer,
                    OD,
                    WT,
                    SMYS,
                    service_years,
                    min_range,
                    M,
                    fatigue_curve,
                    # Dent Information
                    dent_depth, 
                    dent_depth_in, 
                    dent_length, 
                    dent_width, 
                    dent_orientation,
                    # Parameters for Liquid Pipelines
                    Lx,
                    hx,
                    SG,
                    L1,
                    L2,
                    h1,
                    h2,
                    D1,
                    D2,
                    # Dent Strain Analysis
                    eo_max,
                    ei_max,
                    strain_pass_fail,
                    # Remaining Life Analysis
                    0,#SSI,
                    0,#CI,
                    0,#SSI_MD49,
                    0,
                    0,
                    0,
                    SCF,
                    0,#remaining_life_val,
                    str(flag)]
                   
        # =============================================================================
        # Output the Iteration Results
        # =============================================================================

        # Save the results to a storage Excel file
        wbR = load_workbook(wbR_path, data_only=True)
        wbR_sn = wbR.sheetnames
        wbRs = wbR[wbR_sn[0]]
        wbRs.append(results)
        wbR.save(wbR_path)
        wbR.close()
        
        results_data.append(results)
        j += 1