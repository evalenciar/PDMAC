# -*- coding: utf-8 -*-
"""
Dent Strain Analysis v0.1
ASME B31.8-2020 Appendix R Estimating Strain in Dents
Created on 08/25/2022

The purpose of this script is perform a strain analysis on caliper data using
the method in ASME B31.8-Appendix R.

@author: evalencia
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import time
import matplotlib.pyplot as plt
# from openpyxl.formatting.rule import ColorScaleRule

# Update the file_name as needed
file_name = 'Strain Results (v1.0).xlsx'

def calculations(OD, WT, d, L, sd_axial, sd_circ_deg, sd_radius):
    """
    ASME B31.8-2020 Nonmandatory Appendix R Estimating Strain in Dents calculates
    the bending strain in the circumferential direction, e1, the bending strain in
    the longitudinal direction, e2, and the extensional strain in the longitudinal
    direction, e3. 
    
    This function calculates strains e1, e2, and e3 along with the strain for the
    inside and outside pipe surfaces.

    Parameters
    ----------
    OD : float
        the pipe nominal outside diameter, in
    WT: float
        the pipe nominal wall thickness, in
    d : float
        the dent feature depth, in
    L : float
        the dent feature length, in
    circ_window : int
        the smoothing window (number of points to consider) for the circumferential
        smoothing filter. Note: this must be an odd number. Default = 5
    sd_axial : array of floats
        1-D array containing the axial displacement values with shape (axial x 1), in
    sd_circ_deg : array of floats
        1-D array containing the circumferential displacement values with shape (1 x circ), deg
    sd_radius : array of floats
        2-D array containing the radial values with shape (axial x circ), in

    Returns
    -------
    df_eo : DataFrame
        DataFrame containing all of the strain for the outside pipe surface
    df_ei : DataFrame
        DataFrame containing all of the strain for the inside pipe surface
    df_e1 : DataFrame
        DataFrame containing the bending strain in the circumferential direction
    df_e2 : DataFrame
        DataFrame containing the bending strain in the longitudinal direction
    e3 : float
        float value of the extensional strain in the longitudinal direction
    df_R1 : DataFrame
        DataFrame containing the Radius of Curvature in the circumferential plane
    df_R2 : DataFrame
        DataFrame containing the Radius of Curvature in the longitudinal plane
    """
    
    R0 = OD/2
    sd_circ = np.deg2rad(sd_circ_deg)
    
    # Strain calculations
    sd_e1 = np.zeros(sd_radius.shape)
    sd_e2 = np.zeros(sd_radius.shape)

    e3 = (1/2)*(d/L)**2

    # Radius of curvatures
    sd_R1 = np.zeros(sd_radius.shape)
    sd_R2 = np.zeros(sd_radius.shape)
    
    # Calculate the bending strain in the circumferential direction, e1
    for axial_index, circ_profile in enumerate(sd_radius[:,0]):
        circ_profile = sd_radius[axial_index, :]
        # First derivative
        d_circ = np.gradient(circ_profile, sd_circ)
        # d_circ[0] = (circ_profile[1] - circ_profile[-1])/(sd_circ[1] - (2*np.pi - sd_circ[-1]))
        # d_circ[-1] = (circ_profile[-2] - circ_profile[0])/((2*np.pi - sd_circ[-2]) - sd_circ[0])
        # Second derivative
        dd_circ = np.gradient(d_circ, sd_circ)
        # dd_circ[0] = (circ_profile[1] - circ_profile[-1])/(sd_circ[1] - (2*np.pi - sd_circ[-1]))
        # dd_circ[-1] = (circ_profile[-2] - circ_profile[0])/((2*np.pi - sd_circ[-2]) - sd_circ[0])
        # Radius of curvature in polar coordinates
        R1 = (circ_profile**2 + d_circ**2)**(3/2)/abs(circ_profile**2 + 2*d_circ**2 - circ_profile*dd_circ)
        # Calculate e1 and save it for this circumferential profile
        sd_e1[axial_index, :] = (WT/2)*(1/R0 - 1/R1)
        sd_R1[axial_index, :] = R1
        
    # Calculate the bending strain in the longitudinal (axial) direction, e2
    for circ_index, axial_profile in enumerate(sd_radius[0,:]):
        axial_profile = sd_radius[:, circ_index]
        # First derivative
        d_axial = np.gradient(axial_profile, sd_axial)
        # Second derivative
        dd_axial = np.gradient(d_axial, sd_axial)
        # Radius of curvature. Added np.float64 to help division by zero -> inf
        R2 = (1 + d_axial**2)**(3/2)/np.float64(abs(dd_axial))
        R2[R2 == np.inf] = 1000000
        # Calculate e2 and save it for this axial profile
        sd_e2[:, circ_index] = WT/(2*R2)
        sd_R2[:, circ_index] = R2
        
    # Calculate the final strain for the outside pipe, eo, and inside pipe, ei
    ei = (2/np.sqrt(3))*np.sqrt(sd_e1**2 + sd_e1*(sd_e2 + e3) + (sd_e2 + e3)**2)
    eo = (2/np.sqrt(3))*np.sqrt((-sd_e1)**2 + (-sd_e1)*((-sd_e2) + e3) + ((-sd_e2) + e3)**2)
    
    df_ei = pd.DataFrame(data=ei, columns=sd_circ_deg, index=sd_axial)
    df_eo = pd.DataFrame(data=eo, columns=sd_circ_deg, index=sd_axial)
    df_e1 = pd.DataFrame(data=sd_e1, columns=sd_circ_deg, index=sd_axial)
    df_e2 = pd.DataFrame(data=sd_e2, columns=sd_circ_deg, index=sd_axial)
    df_R1 = pd.DataFrame(data=sd_R1, columns=sd_circ_deg, index=sd_axial)
    df_R2 = pd.DataFrame(data=sd_R2, columns=sd_circ_deg, index=sd_axial)
    
    return df_eo, df_ei, df_e1, df_e2, e3, df_R1, df_R2

def export(featureID, df_data, df_eo, df_ei, df_e1, df_e2, e3, df_R1, df_R2):
    """
    ASME B31.8-2020 Nonmandatory Appendix R Estimating Strain in Dents calculates
    the bending strain in the circumferential direction, e1, the bending strain in
    the longitudinal direction, e2, and the extensional strain in the longitudinal
    direction, e3. 
    
    This function exports the calculation results into a standalone Excel workbook
    for further review by the analyst. 

    Parameters
    ----------
    featureID : TYPE
        DESCRIPTION.
    df_data : DataFrame
        DataFrame containing all of the input radius values used for this strain analysis.
    df_eo : DataFrame
        DataFrame containing all of the strain for the outside pipe surface
    df_ei : DataFrame
        DataFrame containing all of the strain for the inside pipe surface
    df_e1 : DataFrame
        DataFrame containing the bending strain in the circumferential direction
    df_e2 : DataFrame
        DataFrame containing the bending strain in the longitudinal direction
    e3 : float
        float value of the extensional strain in the longitudinal direction
    df_R1 : DataFrame
        DataFrame containing the Radius of Curvature in the circumferential plane
    df_R2 : DataFrame
        DataFrame containing the Radius of Curvature in the longitudinal plane

    Returns
    -------
    None, but creates an Excel workbook in the results folder.

    """
    ref_path = templates_path + file_name
    wb = openpyxl.load_workbook(filename=ref_path)
    
    ar_eo = df_eo.to_numpy()
    ar_ei = df_ei.to_numpy()
    ar_e1 = df_e1.to_numpy()
    ar_e2 = df_e2.to_numpy()
    ar_R1 = df_R1.to_numpy()
    ar_R2 = df_R2.to_numpy()
    
    eo_max_index = (np.where(ar_eo == ar_eo.max())[0][0], np.where(ar_eo == ar_eo.max())[1][0])
    ei_max_index = (np.where(ar_ei == ar_ei.max())[0][0], np.where(ar_ei == ar_ei.max())[1][0])
    
    # Save the values into the corresponding sheet in wb
    # The Excel locations will need a row and col adjustment due to the index and header of the DataFrame
    # For the eo and ei Excel Location, need to adjust the column and row number due to the header, empty row, and index column. 
    # Strain conversion
    sc = 100
    wbs = wb['Summary']
    wbs['D2'] = featureID
    wbs['D3'] = sc*ar_eo.max()
    wbs['D4'] = get_column_letter(eo_max_index[1] + 2) + str(eo_max_index[0] + 3) # Excel Location
    wbs['D5'] = df_eo.columns.values[eo_max_index[1]]   # Circ Location
    wbs['D6'] = df_eo.index.values[eo_max_index[0]]     # Axial Location
    wbs['D7'] = sc*ar_ei.max()
    wbs['D8'] = get_column_letter(ei_max_index[1] + 2) + str(ei_max_index[0] + 3) # Excel Location
    wbs['D9'] = df_ei.columns.values[ei_max_index[1]]     # Circ Location
    wbs['D10'] = df_ei.index.values[ei_max_index[0]]  # Axial Location
    wbs['D11'] = sc*ar_e1[eo_max_index] # e1
    wbs['D12'] = sc*ar_e2[eo_max_index] # e2
    wbs['D13'] = sc*e3
    wbs['D14'] = ar_R1[eo_max_index] # R1
    wbs['D15'] = ar_R2[eo_max_index] # R2
    
    # Iterate through all of the DataFrames and save them to the workbook
    print((time_ref + 'Began saving dent strain analysis results to Excel workbook.') % (time.time() - time_start))
    data_names = ['Data','eo','ei','e1','e2','R1','R2']
    data_files = [df_data, df_eo, df_ei, df_e1, df_e2, df_R1, df_R2]
    for i, val in enumerate(data_names):
        wbs = wb[val]
        # for r in dataframe_to_rows(data_files[i], index=True, header=True):
        for r in dataframe_to_rows(data_files[i]):
            wbs.append(r)
        
        # Need to figure out how to remove the extra second row that is empty <- 09/28/2022
        
        # Add formatting here for the data values
        # FUTURE WORK FOR UPDATING THIS SCRIPT 9/12/2023
        
    # Save the resultant Excel workbook into the designated folder
    wb_path = results_path + file_name
    wb.save(filename=wb_path)
    print((time_ref + 'Done. Results available at: ' + wb_path) % (time.time() - time_start))
    
    # Generate a Contour Plot
    plot_files = [df_eo, df_ei, df_e1, df_e2, df_R1, df_R2]
    plot_names = ['eo_Strain_', 'ei_Strain_', 'e1_Strain_', 'e2_Strain_', 'R1_', 'R2_']
    plot_labels = ['Strain (in/in)', 'Strain (in/in)', 'Strain (in/in)', 'Strain (in/in)', 'Radius of Curvature (in)', 'Radius of Curvature (in)']
    plot_titles = ['eo Strain', 'ei Strain', 'e1 Strain', 'e2 Strain', 'Radius of Curvature R1', 'Radius of Curvature R2']
    for i, val in enumerate(plot_files):
        levels = np.linspace(val.values.min(),val.values.max(),17)
        graph_surface = results_path + plot_names[i] + "ID" + str(featureID) + '.png'
        fig1, cp1 = plt.subplots(figsize=(6.85,4), dpi=240)
        cpf1 = cp1.contourf(val.index.values, val.columns.values, val.values.T, levels=levels)
        cbar1 = fig1.colorbar(cpf1)
        cbar1.set_label(plot_labels[i])
        cp1.set_xlabel('Axial Displacement (in)')
        cp1.set_ylabel('Circumferential Orientation (deg)')
        cp1.set_yticks([])
        cp1.set_title("ID" + str(featureID) + ' ' + plot_titles[i])
        fig1.savefig(graph_surface)
    
    # Return the following: Max e_o, Max e_i, 
    return ar_eo.max(), ar_ei.max()

def strain(dent_ID, results_path_og, OD, WT, d, L, df_data, circ_int = 0.5, axial_int = 0.5, time_start_og=0):
    global time_start
    global time_limit
    global templates_path
    global results_path
    global time_ref
    global flag
    
    time_start = time_start_og
    results_path = results_path_og
    time_limit = 60*5
    
    templates_path = 'templates/'
    time_ref = '%03d | '
    flag = []
    
    print((time_ref + '========= STRAIN ===========') % (time.time() - time_start))
    
    # Take the radial, circumferential, and axial data from the input dataframe
    sd_radius = df_data.to_numpy()
    sd_circ = df_data.columns.values
    sd_axial = df_data.index.values
    
    # 07-28-2023 GOAL: Automatically estimate the depth (d) and length (L) of the dent
    # using some parameters similar from MD-4-9. This will only impact e3 which changes
    # the overall strain values ei and eo. Therefore, need to find a way to get a good
    # estimate of the dent shape. Could perhaps use ML for this as well based on shape.
    
    # Calcualte the dent strain
    df_eo, df_ei, df_e1, df_e2, e3, df_R1, df_R2 = calculations(OD, WT, d, L, sd_axial, sd_circ, sd_radius)
    
    # Export the results of the strain analysis
    eo_max, ei_max = export(dent_ID, df_data, df_eo, df_ei, df_e1, df_e2, e3, df_R1, df_R2)
    
    # Check if the strain values exceed the strain limit
    strain_limit = 0.06
    if (eo_max >= strain_limit) or (ei_max >= strain_limit):
        strain_pass_fail = 'Fail'
    elif (eo_max < strain_limit) and (ei_max < strain_limit):
        strain_pass_fail = 'Pass'
    
    return eo_max, ei_max, strain_limit, strain_pass_fail