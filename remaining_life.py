# -*- coding: utf-8 -*-
"""
Remaining Life Analysis v0.1
Created on 10/24/2022

The purpose of this script is to use the dent Stress Concentration Factor (SCF)
and Rainflow Analysis to determine the remaining life.

@author: evalencia
"""

import openpyxl

def export(dent_ID, results_path, OD, WT, grade, service_years, min_pressure, m_exponent, cycles, MD49_bins, SCF, fatigue_curve):
    
    templates_path = 'templates/'
    
    file_name ='Remaining Life.xlsx'
    ref_path = templates_path + file_name
    wb = openpyxl.load_workbook(filename=ref_path)
    
    # Update the values in the Summary tab
    wbs = wb['Summary']
    wbs['D2'] = OD
    wbs['D3'] = WT
    wbs['D4'] = grade
    wbs['D5'] = service_years
    wbs['D6'] = min_pressure
    wbs['D7'] = m_exponent
    wbs['D9'] = SCF
    wbs['D10'] = str(fatigue_curve)
    
    # Import the values in the Rainflow tab
    wbs = wb['Rainflow']
    for row_i, row_val in enumerate(cycles):
        for col_i, cell in enumerate(row_val):
            wbs.cell(row=2 + row_i, column=1 + col_i).value = float(cycles[row_i, col_i])
            
    # Save the resultant Excel workbook into the designated folder
    wb_path = results_path + file_name
    wb.save(filename=wb_path)
    
def export_combined(dent_ID, results_path, OD, WT, grade, min_pressure, m_exponent, SCF, fatigue_curve, T1_data, T2_data):
    
    templates_path = 'templates/'
    
    file_name ='Remaining Life (Combine Damage).xlsx'
    ref_path = templates_path + file_name
    wb = openpyxl.load_workbook(filename=ref_path)
    
    # Update the values in the Summary tab
    wbs = wb['Summary']
    wbs['D2'] = OD
    wbs['D3'] = WT
    wbs['D4'] = grade
    wbs['D5'] = min_pressure
    wbs['D6'] = m_exponent
    wbs['D12'] = SCF
    # wbs['D13'] = str(fatigue_curve)
    
    # Import the values in the Rainflow tab
    T1_service_years, T1_cycles, T1_MD49_bins = T1_data
    T2_service_years, T2_cycles, T2_MD49_bins = T2_data
    
    wbs = wb['T1-Rainflow']
    # Import the Rainflow results
    for row_i, row_val in enumerate(T1_cycles):
        for col_i, cell in enumerate(row_val):
            wbs.cell(row=2 + row_i, column=1 + col_i).value = float(T1_cycles[row_i, col_i])
    # Import the MD-4-9 Bins results
    for row_i, row_val in enumerate(T1_MD49_bins):
        # Start at N3
        wbs.cell(row=3 + row_i, column=14).value = T1_MD49_bins[row_i]
        
    wbs = wb['T2-Rainflow']
    # Import the Rainflow results
    for row_i, row_val in enumerate(T2_cycles):
        for col_i, cell in enumerate(row_val):
            wbs.cell(row=2 + row_i, column=1 + col_i).value = float(T2_cycles[row_i, col_i])
    # Import the MD-4-9 Bins results
    for row_i, row_val in enumerate(T2_MD49_bins):
        # Start at N3
        wbs.cell(row=3 + row_i, column=14).value = T2_MD49_bins[row_i]
            
    # Save the resultant Excel workbook into the designated folder
    wb_path = results_path + file_name
    wb.save(filename=wb_path)