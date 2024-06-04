# -*- coding: utf-8 -*-
"""
Rainflow Analysis
Created on 09/12/2022

The objective of this script is to perform a rainflow analysis on pressure history data.

@author: evalencia
"""

import numpy as np
import pandas as pd
import rainflow
import matplotlib.pyplot as plt
import openpyxl

# =============================================================================
# Functions
# =============================================================================

def Px(Lx,hx,P1,P2,K,L1,L2,h1,h2,D1,D2):
    """
    Taken from API 1176
    
    Note: the version in API 1183 Section 6.6.3.1 Rainflow Counting Equation (5) is incorrect.
    
    Parameters
    ----------
    Lx : float
        the location of point analysis, ft
    hx : float
        the elevation of point analysis, ft
    P1 : array
        the upstream discharge pressure, psig
    P2 : array
        the downstream suction pressure, psig
    K : float
        SG x (0.433 psi/ft), where SG = specific gravity of product
    L1 : float
        the location of upstream discharge station, ft
    L2 : float
        the location of downstream suction station, ft
    h1 : float
        the elevation of upstream discharge station, ft
    h2 : float
        the elevation of downstream suction station, ft
    D1 : float
        the pipe diameter of segment between L1 and Lx, in
    D2 : float
        the pipe diameter of segment between Lx and L2, in

    Returns
    -------
    The intermediate pressure point between pressure sources, psig

    """

    Px = (P1 + K*h1 - P2 - K*h2)*(1/(((Lx - L1)*D2**5)/((L2 - Lx)*D1**5) + 1)) - K*(hx - h2) + P2
    
    return Px

def MD49(cycles, OD, WT, SMYS, service_years, M, min_range=5):
    # Create an empty array for all the MD-4-9 bins
    MD49_bins = np.zeros(28)
    MD49_P_range = np.array([10,20,30,40,50,60,70,
                             10,20,30,40,50,60,
                             10,20,30,40,50,
                             10,20,30,40,
                             10,20,30,
                             10,20,
                             10])
    
    # Remove any MD49_cycles that have a pressure range below the minimum range
    # MD49_cycles[:,2] = [0 for press_range in MD49_cycles[:,0] if press_range < min_range]
    MD49_cycles = cycles.copy()
    for i, val in enumerate(MD49_cycles[:,0]):
        if MD49_cycles[i,0] < min_range: 
            MD49_cycles[i,2] = 0
            
    # MD49_cycles = MD49_cycles[MD49_cycles['Range'] >= min_range]
    
    # Reference Stress Ranges
    SSI_ref_stress = 13000  # psi
    SSI_ref_press = SSI_ref_stress*2*WT/OD
    
    # Convert pressure range into units of % SMYS
    MD49_cycles[:,0] = 100*MD49_cycles[:,0]*OD/(2*WT)/SMYS
    MD49_cycles[:,1] = 100*MD49_cycles[:,1]*OD/(2*WT)/SMYS
    
    # MD49_cycles[:,0] = [(100*press_range*OD/(2*WT)/SMYS) for press_range in MD49_cycles[:,0]]
    # MD49_cycles[:,1] = [(100*press_mean*OD/(2*WT)/SMYS) for press_mean in MD49_cycles[:,1]]
    # MD49_cycles['Range'] = MD49_cycles['Range'].apply(lambda x: 100*x*OD/(2*WT)/SMYS)
    # MD49_cycles['Mean'] = MD49_cycles['Mean'].apply(lambda x: 100*x*OD/(2*WT)/SMYS)
    
    # Iterate through every pressure range cycle
    for i, press_range in enumerate(MD49_cycles[:, 0]):
        # Pressure range: 0 - 10% SMYS
        if MD49_cycles[i, 0] <= 10.0: #if range is 0 - 10% SMYS
            if MD49_cycles[i, 1] <= 20.0: #if mean is 0 - 20% SMYS
                MD49_bins[0] = MD49_bins[0] + MD49_cycles[i, 2] #BIN #1
            elif MD49_cycles[i, 1] <= 30.0: #if mean is 20 - 30% SMYS
                MD49_bins[7] = MD49_bins[7] + MD49_cycles[i, 2] #BIN #8
            elif MD49_cycles[i, 1] <= 40.0: #if mean is 30 - 40% SMYS
                MD49_bins[13] = MD49_bins[13] + MD49_cycles[i, 2] #BIN #14
            elif MD49_cycles[i, 1] <= 50.0: #if mean is 40 - 50% SMYS
                MD49_bins[18] = MD49_bins[18] + MD49_cycles[i, 2] #BIN #19
            elif MD49_cycles[i, 1] <= 60.0: #if mean is 50 - 60% SMYS
                MD49_bins[22] = MD49_bins[22] + MD49_cycles[i, 2] #BIN #23
            elif MD49_cycles[i, 1] <= 70.0: #if mean is 60 - 70% SMYS
                MD49_bins[25] = MD49_bins[25] + MD49_cycles[i, 2] #BIN #26
            else: #if mean is >70% SMYS
                MD49_bins[27] = MD49_bins[27] + MD49_cycles[i, 2] #BIN #28
        # Pressure range: 10 - 20% SMYS
        elif MD49_cycles[i, 0] <= 20.0: #if range is 10 - 20% SMYS
            if MD49_cycles[i, 1] <= 25.0: #if mean is 0 - 25% SMYS
                MD49_bins[1] = MD49_bins[1] + MD49_cycles[i, 2] #BIN #2
            elif MD49_cycles[i, 1] <= 35.0: #if mean is 25 - 35% SMYS
                MD49_bins[8] = MD49_bins[8] + MD49_cycles[i, 2] #BIN #9
            elif MD49_cycles[i, 1] <= 45.0: #if mean is 35 - 45% SMYS
                MD49_bins[14] = MD49_bins[14] + MD49_cycles[i, 2] #BIN #15
            elif MD49_cycles[i, 1] <= 55.0: #if mean is 45 - 55% SMYS
                MD49_bins[19] = MD49_bins[19] + MD49_cycles[i, 2] #BIN #20
            elif MD49_cycles[i, 1] <= 65.0: #if mean is 55 - 65% SMYS
                MD49_bins[23] = MD49_bins[23] + MD49_cycles[i, 2] #BIN #24
            else: #if mean is >65% SMYS
                MD49_bins[26] = MD49_bins[26] + MD49_cycles[i, 2] #BIN #27
        # Pressure range: 20 - 30% SMYS
        elif MD49_cycles[i, 0] <= 30.0: #if range is 20 - 30% SMYS
            if MD49_cycles[i, 1] <= 30.0: #if mean is 0 - 30% SMYS
                MD49_bins[2] = MD49_bins[2] + MD49_cycles[i, 2] #BIN #3
            elif MD49_cycles[i, 1] <= 40.0: #if mean is 30 - 40% SMYS
                MD49_bins[9] = MD49_bins[9] + MD49_cycles[i, 2] #BIN #10
            elif MD49_cycles[i, 1] <= 50.0: #if mean is 40 - 50% SMYS
                MD49_bins[15] = MD49_bins[15] + MD49_cycles[i, 2] #BIN #16
            elif MD49_cycles[i, 1] <= 60.0: #if mean is 50 - 60% SMYS
                MD49_bins[20] = MD49_bins[20] + MD49_cycles[i, 2] #BIN #21
            else: #if mean is >60% SMYS
                MD49_bins[24] = MD49_bins[24] + MD49_cycles[i, 2] #BIN #25
        # Pressure range: 30 - 40% SMYS
        elif MD49_cycles[i, 0] <= 40.0: #if range is 30 - 40% SMYS
            if MD49_cycles[i, 1] <= 35.0: #if mean is 0 - 35% SMYS
                MD49_bins[3] = MD49_bins[3] + MD49_cycles[i, 2] #BIN #4
            elif MD49_cycles[i, 1] <= 45.0: #if mean is 35 - 45% SMYS
                MD49_bins[10] = MD49_bins[10] + MD49_cycles[i, 2] #BIN #11
            elif MD49_cycles[i, 1] <= 55.0: #if mean is 45 - 55% SMYS
                MD49_bins[16] = MD49_bins[16] + MD49_cycles[i, 2] #BIN #17
            else: #if mean is >55% SMYS
                MD49_bins[21] = MD49_bins[21] + MD49_cycles[i, 2] #BIN #22
        # Pressure range: 40 - 50% SMYS
        elif MD49_cycles[i, 0] <= 50.0: #if range is 40 - 50% SMYS
            if MD49_cycles[i, 1] <= 40.0: #if mean is 0 - 40% SMYS
                MD49_bins[4] = MD49_bins[4] + MD49_cycles[i, 2] #BIN #5
            elif MD49_cycles[i, 1] <= 50.0: #if mean is 40 - 50% SMYS
                MD49_bins[11] = MD49_bins[11] + MD49_cycles[i, 2] #BIN #12
            else: #if mean is >50% SMYS
                MD49_bins[17] = MD49_bins[17] + MD49_cycles[i, 2] #BIN #18
        # Pressure range: 50 - 60% SMYS
        elif MD49_cycles[i, 0] <= 60.0: #if range is 50 - 60% SMYS
            if MD49_cycles[i, 1] <= 45.0: #if mean is 0 - 45% SMYS
                MD49_bins[5] = MD49_bins[5] + MD49_cycles[i, 2] #BIN #6
            else: #if mean is >45% SMYS
                MD49_bins[12] = MD49_bins[12] + MD49_cycles[i, 2] #BIN #13
        # Pressure range > 60% SMYS
        else: #if pressure range > 60% SMYS
            MD49_bins[6] = MD49_bins[6] + MD49_cycles[i, 2] #BIN #7
            
    # Calculate the SSI MD49
    # MD49_final_cycles = sum(((MD49_P_range/100)**M)*MD49_bins)/service_years
    # MD49_final_cycles = sum(((((MD49_P_range/100)*SMYS)/SSI_ref_press)**M)*MD49_bins)/service_years
    
    # 07/19/2023 Changed SSI_ref_press to SSI_ref_stress because that is the correct reference
    MD49_final_cycles = sum(((((MD49_P_range/100)*SMYS)/SSI_ref_stress)**M)*MD49_bins)/service_years
    
    # Save the MD49_bins to a .txt file
    np.savetxt('md49_bins.csv',MD49_bins,delimiter=',')
    
    return MD49_final_cycles, MD49_bins

def equivalent_cycles(index, cycles, OD, WT, SMYS, service_years, M, min_range=5):
    """
    Parameters
    ----------
    index : string
        SSI or CI. SSI is the Spectrum Severity Indicator, CI is the Cyclic Index
    cycles : array of floats
        the array output from the rainflow analysis
    service_years : float
        the period of time in years for the pressure history data
    min_range : array
        the threshold value for pressure ranges to consider. Default is 5 psi.

    Returns
    -------
    Either the SSI or CI. 

    """
    equiv_cycles = np.zeros(cycles.shape[0])
    
    # Reference Stress Ranges
    SSI_ref_stress = 13000  # psi
    SSI_ref_press = SSI_ref_stress*2*WT/OD
    CI_ref_stress = 37580   # psi
    CI_ref_press = CI_ref_stress*2*WT/OD
    
    if index.lower() == 'ssi':
        ref_press = SSI_ref_press
    elif index.lower() == 'ci':
        ref_press = CI_ref_press
        
    for i, val in enumerate(cycles[:,0]):
        if cycles[i,0] > min_range: 
            equiv_cycles[i] = ((cycles[i,0]/ref_press)**M)*cycles[i,2]
        else:
            equiv_cycles[i] = 0
            
    num_cycles = sum(equiv_cycles)/service_years
    
    return num_cycles

def liquid_P(press_hist_path, P_cols, press_hist_header, press_hist_skiprows, time_col=0):
    """
    Extract the Upstream Discharge (P1) and Downstream Suction (P2) pressures 
    from the pipeline segment. This function removes any data rows that have 
    either no value or NAN in either P1 or P2.
    
    Parameters
    ----------
    press_hist_path : string
        the path to the pressure history data file.
    P_cols : array of index
        the location of the upstream discharge and downstream suction pressures in the data file. 
    press_hist_header : index
        the data header index. If there is no data header, use None.
    press_hist_skiprows : index
        the number of data rows to skip when importing the data.
    time_col : index
        the location of the time stamp in the pressure history data file.

    Returns
    -------
    P_list : array of floats
        array containing the upstream discharge (P1) and downstream suction (P2) pressures.
    P_time : array of floats
        array containing the time stamp matching the output pressure history.
    """
    
    # Select the file for analysis. Supported files are:
    #   csv, xls, xlsx, xlsm, xlsb, odf, ods, and odt

    # file_path = filedialog.askopenfilename()
    filetypes = ('.xls', '.xlsx', '.xlsm', '.xlsb', '.odf', '.ods', '.odt')
    # pd.set_option('display.max_columns', None)

    if press_hist_path.endswith(filetypes):
        df = pd.read_excel(press_hist_path, index_col=None, header=press_hist_header, skiprows=press_hist_skiprows)
    elif press_hist_path.endswith('.csv'):
        df = pd.read_csv(press_hist_path, index_col=None, header=press_hist_header, skiprows=press_hist_skiprows)
    
    # DataFrame with Pressure History
    df_ph = df
    
    # Treat it as synchronized data. Remove any rows that has either no value or NAN in either P_cols[0] or P_cols[1]
    df_ph.drop(df_ph[(df_ph[df_ph.columns[P_cols[0]]] == None) | (pd.isna(df_ph[df_ph.columns[P_cols[0]]]))].index, inplace=True)
    df_ph.drop(df_ph[(df_ph[df_ph.columns[P_cols[1]]] == None) | (pd.isna(df_ph[df_ph.columns[P_cols[1]]]))].index, inplace=True)
    
    P_time = df_ph[df_ph.columns[time_col]]
    P1 = df_ph[df_ph.columns[P_cols[0]]]
    P2 = df_ph[df_ph.columns[P_cols[1]]]
    
    P_list = [P1, P2]
    
    return P_list, P_time

def liquid(P_list, P_time, results_path, df_segment_dents):
    # Extract the information for the specified dent
    dent_ID, pipe_segment, odometer, OD, WT, SMYS, service_years, min_range, M, fatigue_curve, dent_depth, dent_depth_in, dent_length, dent_width, dent_orientation, Lx, hx, SG, K, L1, L2, h1, h2, D1, D2, flag = df_segment_dents
    
    # Determine the operational pressures at the dent location. Taken from Equation (5) in API 1183 Section 6.6.3.1 Rainflow Counting
    P = Px(Lx, hx, P_list[0], P_list[1], K, L1, L2, h1, h2, D1, D2)
    # Rainflow Analysis using Python package 'rainflow'
    cycles = pd.DataFrame(rainflow.extract_cycles(P)).to_numpy()
    # Calculate the SSI, CI, and SSI MD49
    SSI = equivalent_cycles('ssi', cycles, OD, WT, SMYS, service_years, M, min_range)
    CI = equivalent_cycles('ci', cycles, OD, WT, SMYS, service_years, M, min_range)
    MD49_SSI, MD49_bins = MD49(cycles, OD, WT, SMYS, service_years, M, min_range)
    
    export(dent_ID, results_path, OD, WT, SMYS, service_years, min_range, M, cycles, MD49_bins)
    
    P.to_csv(results_path + 'Pressure_History.csv', header=False, index=False)
    
    # Graphing
    liquid_graphing(P, P_time, results_path)
    
    return SSI, CI, MD49_SSI, cycles, MD49_bins

def liquid_special(T1_P_list, T2_P_list, results_path, df_segment_dents):
    # Extract the information for the specified dent
    dent_ID, pipe_segment, odometer, OD, WT, SMYS, service_years, min_range, M, fatigue_curve, dent_depth, dent_depth_in, dent_length, dent_width, dent_orientation, Lx, hx, SG, K, L1, L2, h1, h2, D1, D2, flag, SCF = df_segment_dents
    
    T1_service_years = 2.5
    T2_service_years = 0.51
    
    # Determine the operational pressures at the dent location. Taken from Equation (5) in API 1183 Section 6.6.3.1 Rainflow Counting
    T1_P = Px(Lx, hx, T1_P_list[0], T1_P_list[1], K, L1, L2, h1, h2, D1, D2)
    T2_P = Px(Lx, hx, T2_P_list[0], T2_P_list[1], K, L1, L2, h1, h2, D1, D2)
    # Rainflow Analysis using Python package 'rainflow'
    T1_cycles = pd.DataFrame(rainflow.extract_cycles(T1_P)).to_numpy()
    T2_cycles = pd.DataFrame(rainflow.extract_cycles(T2_P)).to_numpy()
    # Calculate the SSI, CI, and SSI MD49
    T1_SSI = equivalent_cycles('ssi', T1_cycles, OD, WT, SMYS, T1_service_years, M, min_range)
    T1_CI = equivalent_cycles('ci', T1_cycles, OD, WT, SMYS, T1_service_years, M, min_range)
    T1_SSI_MD49, T1_MD49_bins = MD49(T1_cycles, OD, WT, SMYS, T1_service_years, M, min_range)
    
    T2_SSI = equivalent_cycles('ssi', T2_cycles, OD, WT, SMYS, T2_service_years, M, min_range)
    T2_CI = equivalent_cycles('ci', T2_cycles, OD, WT, SMYS, T2_service_years, M, min_range)
    T2_SSI_MD49, T2_MD49_bins = MD49(T2_cycles, OD, WT, SMYS, T2_service_years, M, min_range)
    
    P = np.concatenate((T1_P, T2_P))
    
    np.savetxt(results_path + 'Pressure_History.csv', P, delimiter=',')
    
    # Graphing
    fig, sp = plt.subplots(figsize=(8,4), dpi=240)
    fig.suptitle('Pressure History for $P_x$', fontsize=16)
    sp.plot(P)
    sp.set_ylabel('Pressure (psi)')
    fig.savefig(results_path + 'pressure_history_Px' + '.png')
    
    return T1_SSI, T1_CI, T1_SSI_MD49, T1_cycles, T1_MD49_bins, T2_SSI, T2_CI, T2_SSI_MD49, T2_cycles, T2_MD49_bins
    

def liquid_graphing(P, P_time, results_path):
    # Save the interpolate pressure history 
    fig, sp = plt.subplots(figsize=(8,4), dpi=240)
    fig.suptitle('Pressure History for $P_x$', fontsize=16)
    sp.scatter(P_time, P, s=0.1)
    sp.set_ylabel('Pressure (psi)')
    sp.set_xlabel('Start Date Time')
    fig.savefig(results_path + 'pressure_history_Px' + '.png')

def gas(press_hist_path, graph_path, P_cols, press_hist_header, press_hist_skiprows, df_segment_dents, time_col=0):
    """
    Compare the pressure history for every station and select the pressure history 
    that results in the most conservative pressure history (i.e., highest SSI).
    
    Parameters
    ----------
    press_hist_path : string
        the path to the pressure history data file.
    P_cols : array of index
        the location of various pressure stations in the flow direction (upstream to downstream).
    press_hist_header : index
        the data header index. If there is no data header, use None.
    press_hist_skiprows : index
        the number of data rows to skip when importing the data.
    df_segment_dents : DataFrame
        DataFrame containing a single dent to calculate the SSI as a reference point for pressure history comparison.
    time_col : index
        the location of the time stamp in the pressure history data file.

    Returns
    -------
    The SSI, CI, and MD49_SSI.
    """
    
    # Extract the information for the specified dent
    dent_ID, pipe_segment, odometer, OD, WT, SMYS, service_years, min_range, M, fatigue_curve, dent_depth, dent_depth_in, dent_length, dent_width, dent_orientation, Lx, hx, SG, K, L1, L2, h1, h2, D1, D2, flag = df_segment_dents
    
    # Select the file for analysis. Supported files are:
    #   csv, xls, xlsx, xlsm, xlsb, odf, ods, and odt

    # file_path = filedialog.askopenfilename()
    filetypes = ('.xls', '.xlsx', '.xlsm', '.xlsb', '.odf', '.ods', '.odt')
    # pd.set_option('display.max_columns', None)

    if press_hist_path.endswith(filetypes):
        df = pd.read_excel(press_hist_path, index_col=None, header=press_hist_header, skiprows=press_hist_skiprows)
    elif press_hist_path.endswith('.csv'):
        df = pd.read_csv(press_hist_path, index_col=None, header=press_hist_header, skiprows=press_hist_skiprows)
    
    # DataFrame with Pressure History
    df_ph = df
    
    # Treat it as individual data. Remove any values that has either no value or NAN in the column
    P_list = []
    cycles_list = []
    P_time_list = []
    SSI_list = []
    CI_list = []
    MD49_SSI_list = []
    # Use the pressure history for every station and compare the end results to find the most conservative value
    for i in range(len(P_cols)):
        # P_list.append(df_ph.drop(df_ph[(df_ph[df_ph.columns[P_cols[i]]] == None) | (pd.isna(df_ph[df_ph.columns[P_cols[i]]]))].index).to_numpy())
        df_ph_temp = df_ph.drop(df_ph[(df_ph[df_ph.columns[P_cols[i]]] == None) | (pd.isna(df_ph[df_ph.columns[P_cols[i]]]))].index)
        # Use the current pressure history to perform rainflow analysis
        # P = P_list[i]
        P_list.append(df_ph_temp.iloc[:,P_cols[i]])
        # Rainflow Analysis using Python package 'rainflow'
        cycles_list.append(pd.DataFrame(rainflow.extract_cycles(P_list[i])).to_numpy())
        # Calculate the SSI, CI, and SSI MD49
        SSI_list.append(equivalent_cycles('ssi', cycles_list[i], OD, WT, SMYS, service_years, M, min_range))
        CI_list.append(equivalent_cycles('ci', cycles_list[i], OD, WT, SMYS, service_years, M, min_range=5))
        MD49_SSI_list.append(MD49(cycles_list[i], OD, WT, SMYS, service_years, M, min_range))
        # Save the Pressure History time for this data
        # P_time_list.append(P_list[time_col])
        P_time_list.append(df_ph_temp.iloc[:,time_col])
        
    # Find the most conservative result (most conservative = highest SSI)
    SSI = max(SSI_list)
    CI = CI_list[SSI_list.index(SSI)]
    MD49_SSI = MD49_SSI_list[SSI_list.index(SSI)][0]
    P_time = P_time_list[SSI_list.index(SSI)]
    cycles = cycles_list[SSI_list.index(SSI)]
    MD49_bins = MD49_SSI_list[SSI_list.index(SSI)][1]
    
    # Graphing
    gas_graphing(P_list, P_time, graph_path)
    
    return SSI, CI, MD49_SSI, cycles, MD49_bins

def gas_graphing(P_list, P_time, results_path):
    # Save the figures to the dent results folder to observe all of the pressure histories
    for i in range(len(P_list)):
        fig, sp = plt.subplots(figsize=(8,4), dpi=240)
        fig.suptitle('Pressure History for P'+str(i), fontsize=16)
        sp.scatter(P_time, P_list[i], s=0.1)
        sp.set_ylabel('Pressure (psi)')
        sp.set_xlabel('Start Date Time')
        fig.savefig(results_path + '_pressure_history_P' + str(i) + '.png')
  
def export(dent_ID, results_path, OD, WT, grade, service_years, min_pressure, m_exponent, cycles, MD49_bins):
    
    templates_path = 'templates/'
    
    file_name = 'Rainflow Results (No History).xlsx'
    ref_path = templates_path + file_name
    wb = openpyxl.load_workbook(filename=ref_path)
    
    # Update the values in the Summary tab
    wbs = wb['Summary']
    wbs['D2'] = OD
    wbs['D3'] = WT
    wbs['D4'] = grade
    wbs['D5'] = service_years
    wbs['D6'] = min_pressure
    wbs['D7'] = m_exponent
    
    # Import the values in the Rainflow tab
    wbs = wb['Rainflow']
    for row_i, row_val in enumerate(cycles):
        for col_i, cell in enumerate(row_val):
            wbs.cell(row=2 + row_i, column=1 + col_i).value = float(cycles[row_i, col_i])
            
    # Import the MD49_bins to the MD49 tab
    wbs = wb['MD49']
    for row_i, row_val in enumerate(MD49_bins):
        wbs.cell(row=3 + row_i, column=5).value = MD49_bins[row_i]
        
    # Save the resultant Excel workbook into the designated folder
    wb_path = results_path + file_name
    wb.save(filename=wb_path)
    
    # Save the MD49_bins alone in a separate file
    txt_path = results_path + 'MD49_bins.csv'
    np.savetxt(fname=txt_path, X=MD49_bins, delimiter=',')